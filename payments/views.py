import uuid
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Sum, Count, Q, F, ExpressionWrapper, FloatField
from django.db.models.functions import ExtractHour
from django.utils import timezone
from django.http import HttpResponse, Http404
import os
import csv
from django.urls import reverse
from datetime import datetime, timedelta
from accounts.decorators import admin_required, staff_or_admin_required, user_required
from .models import Payment, Refund, PaymentLog
from .forms import PaymentMethodForm, GCashPaymentForm, CashPaymentForm, PaymentApprovalForm, RefundRequestForm
from reservations.models import ReservationInvitation
from .stripe_integration import create_checkout_session, create_payment_intent, get_publishable_key
from courts.models import Court
from reservations.models import Reservation, CancellationRequest
from reservations.models import ReservationEquipment
from equipment.models import Equipment
from notifications.models import Notification
from notifications.email_utils import (
    send_payment_confirmed_email,
    send_payment_verification_email,
    send_payment_failed_email,
    send_payment_receipt_email,
    send_refund_confirmed_email,
)
from accounts.models import User


# Human-readable match format labels (e.g. 'mixed_doubles' -> 'Mixed Doubles')
MATCH_FORMAT_LABELS = {
    'singles': 'Singles',
    'doubles': 'Doubles',
    'mixed_doubles': 'Mixed Doubles',
}

# Payment methods that transfer money directly to an organization's own accounts
ONLINE_PAYMENT_METHODS = ('gcash', 'maya', 'bank_transfer')


def _org_payment_context(organization):
    """
    Build the checkout/payment context from an organization's payment settings.
    Payment information is always organization-specific: the court's owner decides
    what is shown and what methods can be used. Falls back to empty values when the
    organization has not configured its payment settings yet.
    """
    ps = getattr(organization, 'payment_settings', None) if organization else None
    if ps is None:
        return {
            'payment_settings': None,
            'payment_configured': False,
            'org_name': organization.name if organization else '',
            'gcash_number': '',
            'gcash_account_name': '',
            'gcash_qr_code_url': '',
            'maya_number': '',
            'maya_account_name': '',
            'bank_name': '',
            'bank_account_name': '',
            'bank_account_number': '',
            'payment_instructions': '',
            'accepted_payment_methods': [],
            'enabled_payment_methods': [],
        }
    return {
        'payment_settings': ps,
        'payment_configured': ps.is_configured and ps.is_active,
        'org_name': organization.name if organization else '',
        'gcash_number': ps.gcash_number or '',
        'gcash_account_name': ps.gcash_account_name or '',
        'gcash_qr_code_url': ps.qr_code.url if ps.qr_code else '',
        'maya_number': ps.maya_number or '',
        'maya_account_name': ps.maya_account_name or '',
        'bank_name': ps.bank_name or '',
        'bank_account_name': ps.bank_account_name or '',
        'bank_account_number': ps.bank_account_number or '',
        'payment_instructions': ps.payment_instructions or '',
        'accepted_payment_methods': ps.accepted_payment_methods or [],
        'enabled_payment_methods': ps.enabled_methods,
    }


def _send_match_invitations(reservation, checkout_data, invited_by):
    """Send match invitations to registered users selected as opponents/teammates."""
    player_ids = []
    for key in ('team2_player1_id', 'team1_player2_id', 'team2_player2_id'):
        val = checkout_data.get(key)
        if val:
            try:
                player_ids.append(int(val))
            except (ValueError, TypeError):
                pass
    
    for uid in player_ids:
        if uid and uid != invited_by.id:
            try:
                invited_user = User.objects.get(id=uid, is_active=True)
                # Create invitation
                ReservationInvitation.objects.get_or_create(
                    reservation=reservation,
                    invited_user=invited_user,
                    defaults={
                        'invited_by': invited_by,
                        'status': 'pending',
                        'message': f"You've been invited to join a match at {reservation.court.name} on {reservation.date}."
                    }
                )
                # Send notification
                Notification.objects.create(
                    user=invited_user,
                    message=f"You've been invited to a match! Reservation #{reservation.id} at {reservation.court.name}."
                )
            except User.DoesNotExist:
                pass



@login_required
@user_required
def checkout_page_view(request, checkout_token):
    """
    Checkout page that takes a session-stored checkout token.
    No reservation is created until the user completes payment.
    """
    checkout_data = request.session.get('checkout_data')
    
    if not checkout_data:
        messages.error(request, 'Your checkout session has expired. Please start a new reservation.')
        return redirect('reservation_create')
    
    court = get_object_or_404(Court, id=checkout_data['court_id'], is_active=True)
    
    # Look up equipment objects and calculate total fee
    equipment_ids = checkout_data.get('equipment_ids', [])
    equipment_objects = list(Equipment.objects.filter(id__in=equipment_ids, is_active=True)) if equipment_ids else []
    total_equipment_fee = sum(float(eq.rental_price) for eq in equipment_objects)
    
    # Reconstruct unbound forms for the template
    gcash_form = GCashPaymentForm()
    cash_form = CashPaymentForm()
    stripe_publishable_key = get_publishable_key()
    
    def _render_checkout(extra_ctx=None):
        """Helper to render the checkout page with common context."""
        total_amount = float(checkout_data['subtotal']) + total_equipment_fee
        raw_format = checkout_data.get('match_format', '') or ''
        match_format_display = MATCH_FORMAT_LABELS.get(raw_format, raw_format.replace('_', ' ').title())
        ctx = {
            'checkout_data': checkout_data,
            'court': court,
            'gcash_form': gcash_form,
            'cash_form': cash_form,
            'payment': None,
            'reservation': None,
            'is_session_checkout': True,
            'equipment_objects': equipment_objects,
            'total_equipment_fee': total_equipment_fee,
            'total_amount': total_amount,
            'stripe_publishable_key': stripe_publishable_key,
            'stripe_enabled': bool(stripe_publishable_key),
            'match_format_display': match_format_display,
            **_org_payment_context(court.organization),
            # Taxes & discounts are placeholders; set to non-zero to display rows
            'tax_amount': 0,
            'discount_amount': 0,
        }
        if extra_ctx:
            ctx.update(extra_ctx)
        return render(request, 'user/payments/checkout.html', ctx)
    
    if request.method == 'POST':
        method = request.POST.get('method')
        
        if not method:
            messages.error(request, 'Please select a payment method.')
            return _render_checkout()

        # Guard: online payment methods require the organization's payment info
        if method in ONLINE_PAYMENT_METHODS:
            org_settings = getattr(court.organization, 'payment_settings', None) if court.organization else None
            if org_settings is None or not org_settings.is_active or method not in org_settings.enabled_methods:
                messages.error(
                    request,
                    'Online payment is not available for this organization yet. '
                    'Please pay at the counter, use card payment, or contact the organization directly.'
                )
                return _render_checkout()
        
        # Re-verify slot availability before creating reservation
        date = datetime.strptime(checkout_data['date'], '%Y-%m-%d').date()
        start_time = datetime.strptime(checkout_data['start_time'], '%H:%M').time()
        end_time = datetime.strptime(checkout_data['end_time'], '%H:%M').time()
        
        overlapping = Reservation.objects.filter(
            court=court,
            date=date,
            status__in=['confirmed', 'pending']
        ).exclude(
            start_time__gte=end_time
        ).exclude(
            end_time__lte=start_time
        )
        
        if overlapping.exists():
            messages.error(request, 'Sorry, this time slot was just booked by another user. Please start a new reservation.')
            request.session.pop('checkout_data', None)
            return redirect('reservation_create')
        
        # Check if slot is in the past
        slot_datetime = datetime.combine(date, start_time)
        if timezone.is_naive(slot_datetime):
            slot_datetime = timezone.make_aware(slot_datetime)
        if slot_datetime < timezone.now():
            messages.error(request, 'This time slot has already passed. Please start a new reservation.')
            request.session.pop('checkout_data', None)
            return redirect('reservation_create')
        
        if method in ONLINE_PAYMENT_METHODS:
            method_label = dict(Payment.METHOD_CHOICES).get(method, method.replace('_', ' ').title())
            gcash_form = GCashPaymentForm(request.POST, request.FILES)
            if gcash_form.is_valid():
                try:
                    with transaction.atomic():
                        # Create the reservation
                        duration_hours = checkout_data['duration_hours']
                        hourly_rate = checkout_data['hourly_rate']
                        subtotal = checkout_data['subtotal']
                        
                        reservation = Reservation.objects.create(
                            user=request.user,
                            court=court,
                            date=date,
                            start_time=start_time,
                            end_time=end_time,
                            duration_hours=duration_hours,
                            hourly_rate=hourly_rate,
                            subtotal=subtotal,
                            equipment_fee=0,
                            total_amount=0,  # Will be recalculated by save()
                            status='pending',
                            notes=checkout_data.get('notes', ''),
                            match_name=checkout_data.get('match_name', ''),
                            match_format=checkout_data.get('match_format', 'singles'),
                            game_type=checkout_data.get('game_type', 'friendly'),
                            scoring_format=checkout_data.get('scoring_format', '11'),
                            points_per_game=checkout_data.get('points_per_game') or 11,
                            games_to_win=checkout_data.get('games_to_win') or 2,
                            win_by_two=checkout_data.get('win_by_two') if checkout_data.get('win_by_two') is not None else True,
                            # Match participants from user selection during reservation
                            team1_player2_id=checkout_data.get('team1_player2_id'),
                            team2_player1_id=checkout_data.get('team2_player1_id'),
                            team2_player2_id=checkout_data.get('team2_player2_id'),
                            team1_name=checkout_data.get('team1_name', ''),
                            team2_name=checkout_data.get('team2_name', ''),
                            guest_players_data=checkout_data.get('guest_players_data'),
                        )
                        # The Reservation.save() method recalculates totals
                        
                        # Attach equipment
                        equipment_fee = 0
                        for eq_id in checkout_data.get('equipment_ids', []):
                            try:
                                equipment = Equipment.objects.get(id=eq_id, is_active=True)
                                if equipment.quantity_available > 0:
                                    ReservationEquipment.objects.create(
                                        reservation=reservation,
                                        equipment=equipment,
                                        quantity=1,
                                        rental_fee=equipment.rental_price
                                    )
                                    equipment_fee += float(equipment.rental_price)
                                    equipment.quantity_available -= 1
                                    equipment.save()
                            except Equipment.DoesNotExist:
                                pass
                        
                        # Update equipment fee and recalculate total
                        reservation.equipment_fee = equipment_fee
                        reservation.total_amount = reservation.calculate_total()
                        reservation.save()
                        
                        # Send invitations to selected registered users
                        _send_match_invitations(reservation, checkout_data, request.user)
                        
                        # Create payment record
                        payment = Payment.objects.create(
                            reservation=reservation,
                            amount=reservation.total_amount,
                            status='pending'
                        )
                        
                        # Update payment with online payment details
                        payment.method = method
                        payment.gcash_reference = gcash_form.cleaned_data['gcash_reference']
                        if 'gcash_proof_image' in request.FILES:
                            payment.gcash_proof_image = request.FILES['gcash_proof_image']
                        payment.save()
                        
                        # Log the payment
                        PaymentLog.objects.create(
                            payment=payment,
                            action=f'{method_label} Payment Submitted',
                            details=f'{method_label} reference: {payment.gcash_reference}',
                            performed_by=request.user
                        )
                        
                        # Notify staff
                        staff_users = User.objects.filter(role__in=['super_admin', 'org_admin', 'org_staff'])
                        for staff in staff_users:
                            Notification.objects.create(
                                user=staff,
                                message=f"New {method_label} payment for reservation #{reservation.id} requires verification."
                            )
                        
                        # Send payment verification email to user
                        send_payment_verification_email(request.user, payment)
                        
                        # Clear session data
                        request.session.pop('checkout_data', None)
                    
                    messages.success(request, f'{method_label} payment details submitted. Please wait for verification.')
                    return redirect('payment_status', payment_id=payment.id)
                    
                except Exception as e:
                    messages.error(request, f'An error occurred while processing your reservation. Please try again. Error: {str(e)}')
                    return _render_checkout()
            else:
                # GCash form invalid
                return _render_checkout()
                
        elif method == 'cash':
            cash_form = CashPaymentForm(request.POST, request.FILES)
            if cash_form.is_valid():
                try:
                    with transaction.atomic():
                        # Create the reservation
                        duration_hours = checkout_data['duration_hours']
                        hourly_rate = checkout_data['hourly_rate']
                        subtotal = checkout_data['subtotal']
                        
                        reservation = Reservation.objects.create(
                            user=request.user,
                            court=court,
                            date=date,
                            start_time=start_time,
                            end_time=end_time,
                            duration_hours=duration_hours,
                            hourly_rate=hourly_rate,
                            subtotal=subtotal,
                            equipment_fee=0,
                            total_amount=0,
                            status='pending',
                            notes=checkout_data.get('notes', ''),
                            match_name=checkout_data.get('match_name', ''),
                            match_format=checkout_data.get('match_format', 'singles'),
                            game_type=checkout_data.get('game_type', 'friendly'),
                            scoring_format=checkout_data.get('scoring_format', '11'),
                            points_per_game=checkout_data.get('points_per_game') or 11,
                            games_to_win=checkout_data.get('games_to_win') or 2,
                            win_by_two=checkout_data.get('win_by_two') if checkout_data.get('win_by_two') is not None else True,
                            # Match participants from user selection during reservation
                            team1_player2_id=checkout_data.get('team1_player2_id'),
                            team2_player1_id=checkout_data.get('team2_player1_id'),
                            team2_player2_id=checkout_data.get('team2_player2_id'),
                            team1_name=checkout_data.get('team1_name', ''),
                            team2_name=checkout_data.get('team2_name', ''),
                            guest_players_data=checkout_data.get('guest_players_data'),
                        )
                        
                        # Attach equipment
                        equipment_fee = 0
                        for eq_id in checkout_data.get('equipment_ids', []):
                            try:
                                equipment = Equipment.objects.get(id=eq_id, is_active=True)
                                if equipment.quantity_available > 0:
                                    ReservationEquipment.objects.create(
                                        reservation=reservation,
                                        equipment=equipment,
                                        quantity=1,
                                        rental_fee=equipment.rental_price
                                    )
                                    equipment_fee += float(equipment.rental_price)
                                    equipment.quantity_available -= 1
                                    equipment.save()
                            except Equipment.DoesNotExist:
                                pass
                        
                        reservation.equipment_fee = equipment_fee
                        reservation.total_amount = reservation.calculate_total()
                        reservation.save()
                        
                        # Send invitations to selected registered users
                        _send_match_invitations(reservation, checkout_data, request.user)
                        
                        # Create payment record
                        payment = Payment.objects.create(
                            reservation=reservation,
                            amount=reservation.total_amount,
                            status='pending'
                        )
                        
                        # Set cash payment method
                        payment.method = 'cash'
                        payment.save()
                        
                        PaymentLog.objects.create(
                            payment=payment,
                            action='Cash Payment Selected',
                            details='User selected cash payment method',
                            performed_by=request.user
                        )
                        
                        # Notify staff
                        staff_users = User.objects.filter(role__in=['super_admin', 'org_admin', 'org_staff'])
                        for staff in staff_users:
                            Notification.objects.create(
                                user=staff,
                                message=f"New cash payment reservation #{reservation.id} requires processing."
                            )
                        
                        # Clear session data
                        request.session.pop('checkout_data', None)
                    
                    messages.success(request, 'Please proceed to the counter to complete your cash payment.')
                    return redirect('payment_status', payment_id=payment.id)
                    
                except Exception as e:
                    messages.error(request, f'An error occurred while processing your reservation. Please try again. Error: {str(e)}')
                    return _render_checkout()
            else:
                return _render_checkout()
            
        elif method == 'stripe':
            # Stripe Checkout - create reservation first, then redirect to Stripe
            try:
                with transaction.atomic():
                    duration_hours = checkout_data['duration_hours']
                    hourly_rate = checkout_data['hourly_rate']
                    subtotal = checkout_data['subtotal']
                    
                    reservation = Reservation.objects.create(
                        user=request.user,
                        court=court,
                        date=date,
                        start_time=start_time,
                        end_time=end_time,
                        duration_hours=duration_hours,
                        hourly_rate=hourly_rate,
                        subtotal=subtotal,
                        equipment_fee=0,
                        total_amount=0,
                        status='pending',
                        notes=checkout_data.get('notes', ''),
                        match_name=checkout_data.get('match_name', ''),
                        match_format=checkout_data.get('match_format', 'singles'),
                        game_type=checkout_data.get('game_type', 'friendly'),
                        scoring_format=checkout_data.get('scoring_format', '11'),
                        points_per_game=checkout_data.get('points_per_game') or 11,
                        games_to_win=checkout_data.get('games_to_win') or 2,
                        win_by_two=checkout_data.get('win_by_two') if checkout_data.get('win_by_two') is not None else True,
                        # Match participants from user selection during reservation
                        team1_player2_id=checkout_data.get('team1_player2_id'),
                        team2_player1_id=checkout_data.get('team2_player1_id'),
                        team2_player2_id=checkout_data.get('team2_player2_id'),
                        team1_name=checkout_data.get('team1_name', ''),
                        team2_name=checkout_data.get('team2_name', ''),
                        guest_players_data=checkout_data.get('guest_players_data'),
                    )
                    
                    equipment_fee = 0
                    for eq_id in checkout_data.get('equipment_ids', []):
                        try:
                            equipment_item = Equipment.objects.get(id=eq_id, is_active=True)
                            if equipment_item.quantity_available > 0:
                                ReservationEquipment.objects.create(
                                    reservation=reservation,
                                    equipment=equipment_item,
                                    quantity=1,
                                    rental_fee=equipment_item.rental_price
                                )
                                equipment_fee += float(equipment_item.rental_price)
                                equipment_item.quantity_available -= 1
                                equipment_item.save()
                        except Equipment.DoesNotExist:
                            pass
                    
                    reservation.equipment_fee = equipment_fee
                    reservation.total_amount = reservation.calculate_total()
                    reservation.save()
                    
                    # Send invitations to selected registered users
                    _send_match_invitations(reservation, checkout_data, request.user)
                    
                    payment = Payment.objects.create(
                        reservation=reservation,
                        amount=reservation.total_amount,
                        status='pending',
                        method='card',
                    )
                    
                    PaymentLog.objects.create(
                        payment=payment,
                        action='Stripe Checkout Initiated',
                        details='User redirected to Stripe Checkout',
                        performed_by=request.user
                    )
                    
                    request.session.pop('checkout_data', None)
                
                # Create Stripe Checkout Session
                success_url = request.build_absolute_uri(
                    reverse('payment_status', args=[payment.id])
                ) + '?stripe_session_id={CHECKOUT_SESSION_ID}'
                cancel_url = request.build_absolute_uri(
                    reverse('reservation_detail', args=[reservation.id])
                )
                
                session = create_checkout_session(payment, success_url, cancel_url)
                
                if session:
                    # Store the session ID
                    payment.stripe_checkout_session_id = session.get('id')
                    payment.save(update_fields=['stripe_checkout_session_id'])
                    return redirect(session.url)
                else:
                    messages.error(request, 'Unable to process Stripe payment. Please try again or choose a different payment method.')
                    return redirect('payment_checkout', reservation_id=reservation.id)
                    
            except Exception as e:
                messages.error(request, f'An error occurred while processing your payment. Please try again. Error: {str(e)}')
                return _render_checkout()
        
        elif method == 'card':
            try:
                with transaction.atomic():
                    duration_hours = checkout_data['duration_hours']
                    hourly_rate = checkout_data['hourly_rate']
                    subtotal = checkout_data['subtotal']
                    
                    reservation = Reservation.objects.create(
                        user=request.user,
                        court=court,
                        date=date,
                        start_time=start_time,
                        end_time=end_time,
                        duration_hours=duration_hours,
                        hourly_rate=hourly_rate,
                        subtotal=subtotal,
                        equipment_fee=0,
                        total_amount=0,
                        status='confirmed',
                        notes=checkout_data.get('notes', ''),
                        match_name=checkout_data.get('match_name', ''),
                        match_format=checkout_data.get('match_format', 'singles'),
                        game_type=checkout_data.get('game_type', 'friendly'),
                        scoring_format=checkout_data.get('scoring_format', '11'),
                        points_per_game=checkout_data.get('points_per_game') or 11,
                        games_to_win=checkout_data.get('games_to_win') or 2,
                        win_by_two=checkout_data.get('win_by_two') if checkout_data.get('win_by_two') is not None else True,
                        # Match participants from user selection during reservation
                        team1_player2_id=checkout_data.get('team1_player2_id'),
                        team2_player1_id=checkout_data.get('team2_player1_id'),
                        team2_player2_id=checkout_data.get('team2_player2_id'),
                        team1_name=checkout_data.get('team1_name', ''),
                        team2_name=checkout_data.get('team2_name', ''),
                        guest_players_data=checkout_data.get('guest_players_data'),
                    )
                    
                    equipment_fee = 0
                    for eq_id in checkout_data.get('equipment_ids', []):
                        try:
                            equipment_item = Equipment.objects.get(id=eq_id, is_active=True)
                            if equipment_item.quantity_available > 0:
                                ReservationEquipment.objects.create(
                                    reservation=reservation,
                                    equipment=equipment_item,
                                    quantity=1,
                                    rental_fee=equipment_item.rental_price
                                )
                                equipment_fee += float(equipment_item.rental_price)
                                equipment_item.quantity_available -= 1
                                equipment_item.save()
                        except Equipment.DoesNotExist:
                            pass
                    
                    reservation.equipment_fee = equipment_fee
                    reservation.total_amount = reservation.calculate_total()
                    reservation.save()
                    
                    # Send invitations to selected registered users
                    _send_match_invitations(reservation, checkout_data, request.user)
                    
                    transaction_id = str(uuid.uuid4())[:8].upper()
                    payment = Payment.objects.create(
                        reservation=reservation,
                        amount=reservation.total_amount,
                        status='paid',
                        method='card',
                        transaction_id=transaction_id,
                    )
                    
                    PaymentLog.objects.create(
                        payment=payment,
                        action='Card Payment Processed',
                        details=f'Transaction ID: {transaction_id}',
                        performed_by=request.user
                    )
                    
                    Notification.objects.create(
                        user=request.user,
                        message=f"Payment successful for reservation #{reservation.id}. Your reservation is confirmed!"
                    )
                    
                    send_payment_confirmed_email(request.user, payment)
                    send_payment_receipt_email(request.user, payment)

                    request.session.pop('checkout_data', None)

                    messages.success(request, 'Payment successful! Your reservation is confirmed.')
                    return redirect('reservation_detail', reservation_id=reservation.id)
                
            except Exception as e:
                messages.error(request, f'An error occurred while processing your payment. Please try again. Error: {str(e)}')
                return _render_checkout()
        
        else:
            messages.error(request, 'Invalid payment method selected.')
            return _render_checkout()
    
    # GET request - show the checkout page with session data
    return _render_checkout()


@login_required
def payment_checkout_view(request, reservation_id):
    # Staff/admin can view any reservation, regular users only their own
    if request.user.is_staff_user() or request.user.is_admin():
        reservation = get_object_or_404(Reservation, id=reservation_id)
    else:
        reservation = get_object_or_404(Reservation, id=reservation_id, user=request.user)
    
    # Check if payment already exists
    try:
        payment = Payment.objects.get(reservation=reservation)
    except Payment.DoesNotExist:
        payment = None
    
    if request.method == 'POST':
        method = request.POST.get('method')

        # Guard: reject unknown/missing payment methods before touching the DB
        if method not in dict(Payment.METHOD_CHOICES):
            messages.error(request, 'Please select a valid payment method.')
            return redirect('payment_checkout', reservation_id=reservation.id)

        # Guard: online payment methods require the organization's payment info
        if method in ONLINE_PAYMENT_METHODS:
            org_settings = getattr(reservation.court.organization, 'payment_settings', None) if reservation.court.organization else None
            if org_settings is None or not org_settings.is_active or method not in org_settings.enabled_methods:
                messages.error(
                    request,
                    'Online payment is not available for this organization yet. '
                    'Please pay at the counter, use card payment, or contact the organization directly.'
                )
                return redirect('payment_checkout', reservation_id=reservation.id)

        if not payment:
            payment = Payment.objects.create(
                reservation=reservation,
                amount=reservation.total_amount,
                status='pending'
            )
        
        # Persist the selected method so the payment status page reflects the choice
        # even when the submitted proof details are invalid and need retrying.
        payment.method = method
        payment.save()

        if method in ONLINE_PAYMENT_METHODS:
            method_label = dict(Payment.METHOD_CHOICES).get(method, method.replace('_', ' ').title())
            gcash_form = GCashPaymentForm(request.POST, request.FILES, instance=payment)
            if gcash_form.is_valid():
                payment = gcash_form.save()
                payment.status = 'pending'
                payment.save()
                
                # Log the payment
                PaymentLog.objects.create(
                    payment=payment,
                    action=f'{method_label} Payment Submitted',
                    details=f'{method_label} reference: {payment.gcash_reference}',
                    performed_by=request.user
                )
                
                # Notify staff
                staff_users = User.objects.filter(role__in=['super_admin', 'org_admin', 'org_staff'])
                for staff in staff_users:
                    Notification.objects.create(
                        user=staff,
                        message=f"New {method_label} payment for reservation #{reservation.id} requires verification."
                    )
                
                # Send payment verification email
                send_payment_verification_email(request.user, payment)
                
                messages.success(request, f'{method_label} payment details submitted. Please wait for verification.')
                return redirect('payment_status', payment_id=payment.id)

            # Invalid online-payment form: re-render checkout with the bound form
            # so the user sees the validation errors instead of a 500 error.
            cash_form = CashPaymentForm(instance=payment) if payment else CashPaymentForm()

        elif method == 'cash':
            cash_form = CashPaymentForm(request.POST, request.FILES, instance=payment)
            if cash_form.is_valid():
                payment = cash_form.save()
                payment.status = 'pending'
                payment.save()

                PaymentLog.objects.create(
                    payment=payment,
                    action='Cash Payment Selected',
                    details='User selected cash payment method',
                    performed_by=request.user
                )

                messages.success(request, 'Please proceed to the counter to complete your cash payment.')
                return redirect('payment_status', payment_id=payment.id)

            # Invalid cash form: re-render checkout with the bound form.
            gcash_form = GCashPaymentForm(instance=payment) if payment else GCashPaymentForm()

        elif method == 'card':
            # Generate transaction ID for card payment
            payment.transaction_id = str(uuid.uuid4())[:8].upper()
            payment.status = 'paid'
            payment.save()
            
            PaymentLog.objects.create(
                payment=payment,
                action='Card Payment Processed',
                details=f'Transaction ID: {payment.transaction_id}',
                performed_by=request.user
            )
            
            # Confirm the reservation
            reservation.status = 'confirmed'
            reservation.save()
            
            Notification.objects.create(
                user=request.user,
                message=f"Payment successful for reservation #{reservation.id}. Your reservation is confirmed!"
            )
            
            # Send payment confirmation and receipt emails
            send_payment_confirmed_email(request.user, payment)
            send_payment_receipt_email(request.user, payment)
            
            messages.success(request, 'Payment successful! Your reservation is confirmed.')
            return redirect('reservation_detail', reservation_id=reservation.id)
    else:
        gcash_form = GCashPaymentForm(instance=payment) if payment else GCashPaymentForm()
        cash_form = CashPaymentForm(instance=payment) if payment else CashPaymentForm()
    
    raw_format = (reservation.match_format or '').lower()
    match_format_display = MATCH_FORMAT_LABELS.get(raw_format, raw_format.replace('_', ' ').title())

    return render(request, 'user/payments/checkout.html', {
        'reservation': reservation,
        'payment': payment,
        'gcash_form': gcash_form,
        'cash_form': cash_form,
        'match_format_display': match_format_display,
        'tax_amount': 0,
        'discount_amount': 0,
        **_org_payment_context(reservation.court.organization),
    })


@login_required
def payment_status_view(request, payment_id):
    payment = get_object_or_404(
        Payment.objects.select_related(
            'reservation',
            'reservation__court',
            'reservation__court__site',
            'reservation__user',
            'reservation__match',
            'cash_received_by',
        ).prefetch_related(
            'reservation__rented_equipment',
            'reservation__rented_equipment__equipment',
        ),
        id=payment_id
    )
    
    # Check permissions
    if payment.reservation.user != request.user and not request.user.is_staff_user() and not request.user.is_admin():
        messages.error(request, 'You do not have permission to view this payment.')
        return redirect('user_dashboard')
    
    # Get payment activity logs
    logs = PaymentLog.objects.filter(payment=payment).select_related('performed_by').order_by('-created_at')
    
    # Get refund details if applicable
    refunds = Refund.objects.filter(payment=payment).select_related(
        'requested_by', 'approved_by'
    ).order_by('-requested_at')
    latest_refund = refunds.first()
    
    return render(request, 'user/payments/payment_status.html', {
        'payment': payment,
        'logs': logs,
        'latest_refund': latest_refund,
        'refunds': refunds,
        # Taxes & discounts are placeholders; set to non-zero to display rows
        'tax_amount': 0,
        'discount_amount': 0,
        **_org_payment_context(payment.reservation.court.organization),
    })


@login_required
@staff_or_admin_required
def staff_payments_view(request):
    
    payments = Payment.objects.all().order_by('-created_at')
    
    # Org-scoping for org_admin and org_staff users
    if request.user.organization:
        payments = payments.filter(reservation__court__organization=request.user.organization)
    
    # Search
    search_query = request.GET.get('search', '').strip()
    if search_query:
        payments = payments.filter(
            Q(id__icontains=search_query) |
            Q(reservation__user__username__icontains=search_query) |
            Q(reservation__user__email__icontains=search_query) |
            Q(reservation__id__icontains=search_query)
        )

    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        payments = payments.filter(status=status_filter)
    
    # Filter by method
    method_filter = request.GET.get('method', '')
    if method_filter:
        payments = payments.filter(method=method_filter)
    
    # Filter by date
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        payments = payments.filter(created_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(created_at__date__lte=date_to)
    
    # Sorting
    sort_by = request.GET.get('sort_by', '-created_at')
    sort_order = request.GET.get('sort_order', 'desc')
    allowed_sort_fields = ['id', 'amount', 'method', 'status', 'created_at']
    if sort_by.lstrip('-') in allowed_sort_fields:
        if sort_order == 'asc' and sort_by.startswith('-'):
            sort_by = sort_by[1:]
        elif sort_order == 'desc' and not sort_by.startswith('-'):
            sort_by = '-' + sort_by
        payments = payments.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(payments, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Statistics (scoped for org)
    base_paid = Payment.objects.filter(status='paid')
    base_pending = Payment.objects.filter(status='pending')
    if request.user.organization:
        org_filter = Q(reservation__court__organization=request.user.organization)
        base_paid = base_paid.filter(org_filter)
        base_pending = base_pending.filter(org_filter)
    
    total_paid = base_paid.aggregate(Sum('amount'))['amount__sum'] or 0
    total_pending = base_pending.aggregate(Sum('amount'))['amount__sum'] or 0
    
    today_paid = base_paid.filter(
        created_at__date=timezone.now().date()
    )
    today_revenue = today_paid.aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Counts for summary cards
    paid_count = base_paid.count()
    pending_count = base_pending.count()
    refunded_count = Payment.objects.filter(status='refunded').count()
    if request.user.organization:
        refunded_count = Payment.objects.filter(status='refunded', reservation__court__organization=request.user.organization).count()
    total_payments = base_paid.count() + base_pending.count()
    
    return render(request, 'staff/payments/staff_payments.html', {
        'payments': page_obj.object_list,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'search_query': search_query,
        'status_filter': status_filter,
        'method_filter': method_filter,
        'date_from': date_from,
        'date_to': date_to,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'today_revenue': today_revenue,
        'paid_count': paid_count,
        'pending_count': pending_count,
        'refunded_count': refunded_count,
        'total_payments': total_payments,
    })


@login_required
@staff_or_admin_required
def verify_payment_view(request, payment_id):

    payment_qs = Payment.objects.select_related(
        'reservation',
        'reservation__court',
        'reservation__court__site',
        'reservation__user',
        'cash_received_by',
    ).prefetch_related(
        'reservation__rented_equipment',
        'reservation__rented_equipment__equipment',
    )

    # Org-scoping for org_admin and org_staff
    if request.user.organization:
        payment_qs = payment_qs.filter(reservation__court__organization=request.user.organization)

    payment = get_object_or_404(payment_qs, id=payment_id)

    # Determine template and redirect URL based on user role
    if request.user.is_org_admin():
        template_name = 'admin/payments/verify_payment.html'
        redirect_url = 'org_admin_payments'
    else:
        template_name = 'staff/payments/verify_payment.html'
        redirect_url = 'staff_payments'

    # Get payment activity logs
    logs = PaymentLog.objects.filter(payment=payment).select_related('performed_by').order_by('-created_at')

    # Get refund details if applicable
    refunds = Refund.objects.filter(payment=payment).select_related(
        'requested_by', 'approved_by'
    ).order_by('-requested_at')
    latest_refund = refunds.first()

    if request.method == 'POST':
        form = PaymentApprovalForm(request.POST, instance=payment)
        if form.is_valid():
            old_status = payment.status
            payment = form.save()

            # Update reservation status if payment is verified
            if payment.status == 'paid' and old_status != 'paid':
                payment.reservation.status = 'confirmed'
                payment.reservation.save()

                # Set cash payment details
                if payment.method == 'cash':
                    payment.cash_received_by = request.user
                    payment.cash_received_at = timezone.now()
                    payment.save()

                # Notify user
                Notification.objects.create(
                    user=payment.reservation.user,
                    message=f"Your payment for reservation #{payment.reservation.id} has been verified. Your reservation is confirmed!"
                )
                send_payment_confirmed_email(payment.reservation.user, payment)

            # Log the action
            PaymentLog.objects.create(
                payment=payment,
                action=f'Payment {payment.status}',
                details=f'Payment status changed from {old_status} to {payment.status}',
                performed_by=request.user
            )

            messages.success(request, f'Payment #{payment.id} has been {payment.status}.')
            return redirect(redirect_url)
    else:
        form = PaymentApprovalForm(instance=payment)

    return render(request, template_name, {
        'form': form,
        'payment': payment,
        'logs': logs,
        'latest_refund': latest_refund,
        'refunds': refunds,
    })


@login_required
@staff_or_admin_required
def cash_payment_confirmation_view(request, payment_id):
    """
    Dedicated cash payment counter confirmation page for staff.
    Provides a streamlined, polished interface for staff to mark cash payments
    as received and confirm the reservation in one action.
    """
    payment = get_object_or_404(Payment, id=payment_id)

    # Ensure it's a cash payment and still pending
    if payment.method != 'cash':
        messages.error(request, 'This is not a cash payment.')
        return redirect('staff_payments')

    if payment.status != 'pending':
        messages.error(request, 'This payment has already been processed.')
        return redirect('staff_payments')

    # Check org-scoping
    if request.user.organization and payment.reservation.court.organization != request.user.organization:
        messages.error(request, 'You do not have permission to process this payment.')
        return redirect('staff_payments')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'confirm':
            old_status = payment.status

            # Mark payment as paid
            payment.status = 'paid'
            payment.cash_received_by = request.user
            payment.cash_received_at = timezone.now()
            payment.transaction_id = f"CASH-{payment.id}-{timezone.now().strftime('%Y%m%d%H%M')}"
            payment.save()

            # Confirm the reservation
            reservation = payment.reservation
            reservation.status = 'confirmed'
            reservation.save()

            # Log the action
            PaymentLog.objects.create(
                payment=payment,
                action='Cash Payment Confirmed at Counter',
                details=f'Cash payment of ₱{payment.amount} received by {request.user.get_full_name() or request.user.username}. Reservation #{reservation.id} confirmed.',
                performed_by=request.user
            )

            # Notify the user
            Notification.objects.create(
                user=reservation.user,
                message=f"Your cash payment for reservation #{reservation.id} has been received and confirmed at the counter. Your reservation is confirmed!"
            )
            send_payment_confirmed_email(reservation.user, payment)

            messages.success(
                request,
                f'Cash payment of ₱{payment.amount} received successfully! Reservation #{reservation.id} is now confirmed.'
            )
            return redirect('staff_payments')

        elif action == 'reject':
            # Mark as failed/cancelled
            payment.status = 'failed'
            payment.save()

            PaymentLog.objects.create(
                payment=payment,
                action='Cash Payment Rejected',
                details=f'Cash payment rejected by {request.user.get_full_name() or request.user.username}.',
                performed_by=request.user
            )

            # Send payment failed notification
            send_payment_failed_email(
                payment.reservation.user,
                payment,
                'Cash payment rejected at the counter.'
            )
            messages.warning(request, f'Cash payment #{payment.id} has been marked as not received.')
            return redirect('staff_payments')

    return render(request, 'staff/payments/cash_confirmation.html', {
        'payment': payment,
        'reservation': payment.reservation,
    })


@login_required
def payment_receipt_view(request, payment_id):
    """View/print an official receipt for a payment."""
    payment = get_object_or_404(
        Payment.objects.select_related(
            'reservation', 'reservation__court', 'reservation__user',
            'cash_received_by',
        ).prefetch_related(
            'reservation__rented_equipment__equipment',
            'logs__performed_by',
        ),
        id=payment_id
    )
    
    # Check permissions
    is_owner = payment.reservation.user == request.user
    is_staff = request.user.is_staff_user() or request.user.is_admin()
    if not is_owner and not is_staff:
        messages.error(request, 'You do not have permission to view this receipt.')
        return redirect('user_dashboard')
    
    # Get payment logs
    logs = PaymentLog.objects.filter(payment=payment).select_related('performed_by').order_by('-created_at')
    
    return render(request, 'user/payments/receipt.html', {
        'payment': payment,
        'logs': logs,
    })


@login_required
@user_required
def payment_history_view(request):
    payments = Payment.objects.filter(reservation__user=request.user).order_by('-created_at')
    
    # Compute stats
    total_spent = payments.filter(status='paid').aggregate(Sum('amount'))['amount__sum'] or 0
    paid_count = payments.filter(status='paid').count()
    pending_count = payments.filter(status='pending').count()
    refunded_count = payments.filter(status='refunded').count()
    total_count = payments.count()
    
    return render(request, 'user/payments/payment_history.html', {
        'payments': payments,
        'total_spent': total_spent,
        'paid_count': paid_count,
        'pending_count': pending_count,
        'refunded_count': refunded_count,
        'total_count': total_count,
    })


@login_required
def view_payment_proof_view(request, payment_id):
    """View the GCash payment proof image"""
    payment = get_object_or_404(Payment, id=payment_id)

    # Check permissions - user can view their own, staff/admin can view all
    if payment.reservation.user != request.user and not request.user.is_staff_user() and not request.user.is_admin():
        messages.error(request, 'You do not have permission to view this payment proof.')
        return redirect('payment_status', payment_id=payment.id)

    # Check if image exists
    if not payment.gcash_proof_image:
        messages.error(request, 'No payment proof image available for this payment.')
        return redirect('payment_status', payment_id=payment.id)

    return render(request, 'admin/payments/view_proof.html', {
        'payment': payment,
        'proof_image': payment.gcash_proof_image
    })


@login_required
def serve_payment_proof_image(request, payment_id):
    """Serve the payment proof image directly"""
    payment = get_object_or_404(Payment, id=payment_id)

    # Check permissions - user can view their own, staff/admin can view all
    if payment.reservation.user != request.user and not request.user.is_staff_user() and not request.user.is_admin():
        raise Http404("Payment proof not found")

    # Check if image exists
    if not payment.gcash_proof_image or not payment.gcash_proof_image.name:
        raise Http404("Payment proof image not found")

    try:
        # Serve the image file
        with open(payment.gcash_proof_image.path, 'rb') as f:
            image_data = f.read()

        # Determine content type based on file extension
        content_type = 'image/jpeg'
        if payment.gcash_proof_image.name.lower().endswith('.png'):
            content_type = 'image/png'
        elif payment.gcash_proof_image.name.lower().endswith('.gif'):
            content_type = 'image/gif'
        elif payment.gcash_proof_image.name.lower().endswith('.webp'):
            content_type = 'image/webp'

        response = HttpResponse(image_data, content_type=content_type)
        return response
    except FileNotFoundError:
        raise Http404("Payment proof image file not found")
    except Exception as e:
        raise Http404("Error loading payment proof image")


@login_required
@admin_required
def admin_payments_view(request):
    """Admin-only payment management with enhanced features"""

    payments = Payment.objects.all().order_by('-created_at')
    
    # Org-scoping for org_admin users
    if request.user.is_org_admin() and request.user.organization:
        payments = payments.filter(reservation__court__organization=request.user.organization)

    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        payments = payments.filter(status=status_filter)

    # Filter by method
    method_filter = request.GET.get('method', '')
    if method_filter:
        payments = payments.filter(method=method_filter)

    # Filter by date range
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        payments = payments.filter(created_at__date__gte=date_from)
    if date_to:
        payments = payments.filter(created_at__date__lte=date_to)

    # Filter by user role (who made the payment verification)
    verified_by_filter = request.GET.get('verified_by', '')
    if verified_by_filter:
        if verified_by_filter == 'admin':
            payments = payments.filter(cash_received_by__role='admin')
        elif verified_by_filter == 'staff':
            payments = payments.filter(cash_received_by__role='staff')
        elif verified_by_filter == 'system':
            payments = payments.filter(cash_received_by__isnull=True, method='card')

    # Enhanced Statistics
    today = timezone.now().date()

    # Today's stats
    today_payments = Payment.objects.filter(created_at__date=today)
    today_revenue = today_payments.filter(status='paid').aggregate(Sum('amount'))['amount__sum'] or 0
    today_count = today_payments.filter(status='paid').count()

    # This week's stats
    week_start = today - timedelta(days=today.weekday())
    week_payments = Payment.objects.filter(created_at__date__gte=week_start)
    week_revenue = week_payments.filter(status='paid').aggregate(Sum('amount'))['amount__sum'] or 0
    week_count = week_payments.filter(status='paid').count()

    # This month's stats
    month_start = today.replace(day=1)
    month_payments = Payment.objects.filter(created_at__date__gte=month_start)
    month_revenue = month_payments.filter(status='paid').aggregate(Sum('amount'))['amount__sum'] or 0
    month_count = month_payments.filter(status='paid').count()

    # Overall totals
    total_paid = Payment.objects.filter(status='paid').aggregate(Sum('amount'))['amount__sum'] or 0
    total_pending = Payment.objects.filter(status='pending').aggregate(Sum('amount'))['amount__sum'] or 0
    total_refunded = Payment.objects.filter(status='refunded').aggregate(Sum('amount'))['amount__sum'] or 0

    # Payment method breakdown
    method_breakdown = Payment.objects.filter(status='paid').values('method').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')

    # Verification stats - who processed payments
    verification_stats = Payment.objects.filter(
        status='paid',
        cash_received_by__isnull=False
    ).values('cash_received_by__role').annotate(
        total=Sum('amount'),
        count=Count('id')
    )

    # Pending verifications (online payments pending)
    pending_gcash = Payment.objects.filter(status='pending', method__in=['gcash', 'maya', 'bank_transfer']).count()
    pending_cash = Payment.objects.filter(status='pending', method='cash').count()

    # Sorting
    sort_by = request.GET.get('sort_by', '-created_at')
    sort_order = request.GET.get('sort_order', 'desc')
    allowed_sort_fields = ['id', 'amount', 'method', 'status', 'created_at']
    if sort_by.lstrip('-') in allowed_sort_fields:
        if sort_order == 'asc' and sort_by.startswith('-'):
            sort_by = sort_by[1:]
        elif sort_order == 'desc' and not sort_by.startswith('-'):
            sort_by = '-' + sort_by
        payments = payments.order_by(sort_by)

    # Pagination
    paginator = Paginator(payments, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Recent payment logs
    recent_logs = PaymentLog.objects.select_related('payment', 'performed_by').order_by('-created_at')[:10]

    context = {
        'payments': page_obj.object_list,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'sort_by': sort_by,
        'sort_order': sort_order,
        'status_filter': status_filter,
        'method_filter': method_filter,
        'date_from': date_from,
        'date_to': date_to,
        'verified_by_filter': verified_by_filter,
        # Stats
        'today_revenue': today_revenue,
        'today_count': today_count,
        'week_revenue': week_revenue,
        'week_count': week_count,
        'month_revenue': month_revenue,
        'month_count': month_count,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'total_refunded': total_refunded,
        # Breakdowns
        'method_breakdown': method_breakdown,
        'verification_stats': verification_stats,
        'pending_gcash': pending_gcash,
        'pending_cash': pending_cash,
        'recent_logs': recent_logs,
    }

    return render(request, 'admin/payments/admin_payments.html', context)


@login_required
@admin_required
def admin_cancellation_refunds_view(request):
    """Admin page for reviewing cancellation refund details and marking payments refunded."""

    cancellations = CancellationRequest.objects.select_related(
        'reservation',
        'reservation__user',
        'reservation__court',
        'requested_by',
    ).order_by('-requested_at')

    if request.method == 'POST':
        cancellation_id = request.POST.get('cancellation_id')
        cancellation = get_object_or_404(cancellations, id=cancellation_id)

        try:
            payment = cancellation.reservation.payment
        except Payment.DoesNotExist:
            messages.error(request, 'No payment record was found for this cancellation.')
            return redirect('super_admin_cancellation_refunds')

        old_status = payment.status
        original_amount = payment.amount
        refund_amount = original_amount - cancellation.deduction_amount

        now = timezone.now()

        payment.status = 'refunded'
        payment.save(update_fields=['status', 'updated_at'])

        cancellation.refund_processed = True
        cancellation.refund_processed_at = now
        cancellation.save(update_fields=['refund_processed', 'refund_processed_at'])

        refund = Refund.objects.filter(payment=payment).order_by('-requested_at').first()
        created = refund is None
        if created:
            refund = Refund(payment=payment, requested_by=cancellation.requested_by)

        refund.amount = refund_amount
        refund.reason = f"{cancellation.reason} ({cancellation.deduction_percentage}% cancellation fee applied)"
        refund.status = 'processed'
        refund.approved_by = request.user
        refund.approved_at = now
        refund.processed_at = now
        refund.save()

        PaymentLog.objects.create(
            payment=payment,
            action='Payment Refunded',
            details=f'Payment status changed from {old_status} to refunded via Cancellation & Refund page. Refund record #{refund.id} {"created" if created else "updated"}.',
            performed_by=request.user
        )
        
        Notification.objects.create(
            user=payment.reservation.user,
            message=f"Your payment for reservation #{payment.reservation.id} has been marked as refunded."
        )
        # Send refund email
        try:
            cancellation = CancellationRequest.objects.filter(reservation=payment.reservation).first()
            if cancellation:
                send_refund_confirmed_email(payment.reservation.user, cancellation)
        except Exception:
            pass

        messages.success(request, f'Payment #{payment.id} has been updated to Refunded.')
        return redirect('super_admin_cancellation_refunds')

    refund_rows = []
    for cancellation in cancellations:
        payment = getattr(cancellation.reservation, 'payment', None)
        original_amount = payment.amount if payment else cancellation.reservation.total_amount
        refund_amount = original_amount - cancellation.deduction_amount
        refund_rows.append({
            'cancellation': cancellation,
            'payment': payment,
            'original_amount': original_amount,
            'refund_amount': refund_amount,
        })

    return render(request, 'admin/payments/cancellation_refunds.html', {
        'refund_rows': refund_rows,
    })


@login_required
@admin_required
def revenue_report_view(request):

    # Date range
    date_from = request.GET.get('date_from', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))
    court_filter = request.GET.get('court', '')
    method_filter = request.GET.get('method', '')

    # Base payments queryset
    payments = Payment.objects.filter(
        status='paid',
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    ).select_related('reservation', 'reservation__court', 'reservation__user')
    
    # Org-scoping for org_admin users
    if request.user.is_org_admin() and request.user.organization:
        payments = payments.filter(reservation__court__organization=request.user.organization)

    if court_filter:
        payments = payments.filter(reservation__court_id=court_filter)
    if method_filter:
        payments = payments.filter(method=method_filter)

    today = timezone.now().date()
    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
    days_in_range = (date_to_obj - date_from_obj).days + 1

    # ==================== 1. SUMMARY METRICS ====================

    # Total revenue for selected period
    total_revenue = payments.aggregate(Sum('amount'))['amount__sum'] or 0
    total_paid_bookings = payments.count()
    avg_revenue_per_booking = total_revenue / total_paid_bookings if total_paid_bookings > 0 else 0

    # Daily revenue
    daily_revenue = payments.count()
    avg_revenue_per_day = total_revenue / days_in_range if days_in_range > 0 else 0

    # Org-scoped base paid queryset for all downstream stats
    base_paid_qs = Payment.objects.filter(status='paid')
    if request.user.is_org_admin() and request.user.organization:
        base_paid_qs = base_paid_qs.filter(reservation__court__organization=request.user.organization)

    # Weekly revenue (current week)
    week_start = today - timedelta(days=today.weekday())
    weekly_revenue = base_paid_qs.filter(
        created_at__date__gte=week_start,
        created_at__date__lte=today
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    # Monthly revenue (current month)
    month_start = today.replace(day=1)
    monthly_revenue = base_paid_qs.filter(
        created_at__date__gte=month_start,
        created_at__date__lte=today
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    # Yearly revenue (current year)
    year_start = today.replace(month=1, day=1)
    yearly_revenue = base_paid_qs.filter(
        created_at__date__gte=year_start,
        created_at__date__lte=today
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    # Revenue growth comparison (previous period)
    prev_date_from = date_from_obj - timedelta(days=days_in_range)
    prev_date_to = date_from_obj - timedelta(days=1)
    prev_payments = base_paid_qs.filter(
        created_at__date__gte=prev_date_from,
        created_at__date__lte=prev_date_to
    )
    prev_revenue = prev_payments.aggregate(Sum('amount'))['amount__sum'] or 0
    revenue_growth = ((total_revenue - prev_revenue) / prev_revenue * 100) if prev_revenue > 0 else 0

    # Booking growth
    prev_bookings = prev_payments.count()
    booking_growth = ((total_paid_bookings - prev_bookings) / prev_bookings * 100) if prev_bookings > 0 else 0

    summary_metrics = {
        'total_revenue': total_revenue,
        'daily_revenue': daily_revenue,
        'weekly_revenue': weekly_revenue,
        'monthly_revenue': monthly_revenue,
        'yearly_revenue': yearly_revenue,
        'total_paid_bookings': total_paid_bookings,
        'avg_revenue_per_booking': avg_revenue_per_booking,
        'avg_revenue_per_day': avg_revenue_per_day,
        'revenue_growth': revenue_growth,
        'booking_growth': booking_growth,
        'prev_revenue': prev_revenue,
        'prev_bookings': prev_bookings,
    }

    # ==================== 2. REVENUE BY SOURCE ====================

    # Court reservations (hourly rentals)
    court_reservations = payments.filter(
        reservation__match_name__isnull=True
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Game/match fees (organized play, tournaments)
    game_fees = payments.filter(
        reservation__match_name__isnull=False
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Equipment rentals
    equipment_revenue = payments.filter(
        reservation__equipment_fee__gt=0
    ).aggregate(
        total=Sum('reservation__equipment_fee')
    )['total'] or 0

    # Court fees only (subtotal)
    court_fees_only = payments.aggregate(
        total=Sum('reservation__subtotal')
    )['total'] or 0

    revenue_by_source = {
        'court_reservations': court_reservations,
        'game_fees': game_fees,
        'equipment_rentals': equipment_revenue,
        'court_fees_only': court_fees_only,
    }

    # ==================== 3. BOOKING DETAILS ====================

    booking_details = payments.select_related('reservation', 'reservation__court').prefetch_related(
        'reservation__rented_equipment', 'reservation__rented_equipment__equipment'
    ).order_by('-created_at')

    # ==================== 4. PAYMENT INFORMATION ====================

    # Revenue by payment method
    revenue_by_method_list = list(payments.values('method').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total'))

    # Calculate percentages
    total_rev = summary_metrics['total_revenue']
    for item in revenue_by_method_list:
        item['percentage'] = (float(item['total'] or 0) / float(total_rev) * 100) if total_rev > 0 else 0

    # Org-scoped base all-status queryset for breakdown stats
    base_all_qs = Payment.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    )
    if request.user.is_org_admin() and request.user.organization:
        base_all_qs = base_all_qs.filter(reservation__court__organization=request.user.organization)

    # Payment status breakdown
    payment_status_breakdown = base_all_qs.values('status').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')

    # Pending payments
    pending_payments = base_all_qs.filter(
        status='pending'
    ).aggregate(total=Sum('amount'))['total'] or 0

    payment_info = {
        'revenue_by_method': revenue_by_method_list,
        'status_breakdown': payment_status_breakdown,
        'pending_amount': pending_payments,
    }

    # ==================== 5. DISCOUNTS AND PROMOTIONS ====================

    # Calculate potential discounts (difference between standard rate and actual)
    standard_rate_revenue = payments.aggregate(
        total=Sum(ExpressionWrapper(
            F('reservation__court__hourly_rate') * F('reservation__duration_hours'),
            output_field=FloatField()
        ))
    )['total'] or 0

    # Off-peak analysis (early morning and late evening bookings)
    off_peak_bookings = payments.filter(
        Q(reservation__start_time__hour__lt=8) | Q(reservation__start_time__hour__gte=18)
    ).count()

    discounts_data = {
        'off_peak_bookings': off_peak_bookings,
    }

    # ==================== 6. REFUNDS AND CANCELLATIONS ====================

    # Refunds processed
    refunds = Refund.objects.filter(
        status='processed',
        processed_at__date__gte=date_from,
        processed_at__date__lte=date_to
    )
    if request.user.is_org_admin() and request.user.organization:
        refunds = refunds.filter(payment__reservation__court__organization=request.user.organization)
    total_refunds = refunds.aggregate(total=Sum('amount'))['total'] or 0
    refund_count = refunds.count()

    # Cancelled reservations (no-show or cancelled)
    cancelled_base_qs = Payment.objects.filter(
        reservation__status__in=['cancelled', 'rejected'],
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    )
    if request.user.is_org_admin() and request.user.organization:
        cancelled_base_qs = cancelled_base_qs.filter(reservation__court__organization=request.user.organization)
    cancelled_revenue = cancelled_base_qs.aggregate(total=Sum('amount'))['total'] or 0
    cancelled_count = cancelled_base_qs.count()

    refunds_data = {
        'total_refunds': total_refunds,
        'refund_count': refund_count,
        'cancelled_revenue': cancelled_revenue,
        'cancelled_count': cancelled_count,
        'net_revenue': total_revenue - total_refunds,
    }

    # ==================== 7. CHARTS DATA ====================

    # Daily revenue for chart
    daily_revenue_chart = payments.extra(
        select={'date': 'DATE(payments.created_at)'}
    ).values('date').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('date')

    # Revenue by court
    revenue_by_court = payments.values('reservation__court__name').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')

    # Hourly distribution - use values from related reservation
    hourly_distribution = payments.annotate(
        hour=ExtractHour('reservation__start_time')
    ).values('hour').annotate(
        count=Count('id'),
        revenue=Sum('amount')
    ).order_by('hour')

    # Prepare chart data for JSON serialization
    chart_data = {
        'daily_labels': [item['date'].strftime('%b %d') if hasattr(item['date'], 'strftime') else str(item['date'])[:6] for item in daily_revenue_chart],
        'daily_revenue': [float(item['total'] or 0) for item in daily_revenue_chart],
        'daily_bookings': [item['count'] for item in daily_revenue_chart],
        'court_labels': [item['reservation__court__name'] or 'Unknown' for item in revenue_by_court],
        'court_revenue': [float(item['total'] or 0) for item in revenue_by_court],
        'hourly_labels': [f"{int(item['hour'] or 0)}:00" for item in hourly_distribution],
        'hourly_bookings': [item['count'] for item in hourly_distribution],
        'hourly_revenue': [float(item['revenue'] or 0) for item in hourly_distribution],
    }

    # Court list for filter
    courts = Court.objects.filter(is_active=True)

    return render(request, 'admin/payments/revenue_report.html', {
        'date_from': date_from,
        'date_to': date_to,
        'court_filter': court_filter,
        'method_filter': method_filter,
        'summary_metrics': summary_metrics,
        'revenue_by_source': revenue_by_source,
        'booking_details': booking_details,
        'payment_info': payment_info,
        'discounts_data': discounts_data,
        'refunds_data': refunds_data,
        'daily_revenue_chart': daily_revenue_chart,
        'revenue_by_court': revenue_by_court,
        'hourly_distribution': hourly_distribution,
        'chart_data': chart_data,
        'courts': courts,
    })


@login_required
@admin_required
def revenue_report_export_view(request):
    """Export revenue report as CSV or Excel (.xlsx)"""

    # Get filters
    date_from = request.GET.get('date_from', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))
    court_filter = request.GET.get('court', '')
    method_filter = request.GET.get('method', '')
    export_format = request.GET.get('format', 'csv')

    # Base payments queryset
    payments = Payment.objects.filter(
        status='paid',
        created_at__date__gte=date_from,
        created_at__date__lte=date_to
    ).select_related('reservation', 'reservation__court', 'reservation__user')
    
    # Org-scoping for org_admin users
    if request.user.is_org_admin() and request.user.organization:
        payments = payments.filter(reservation__court__organization=request.user.organization)

    if court_filter:
        payments = payments.filter(reservation__court_id=court_filter)
    if method_filter:
        payments = payments.filter(method=method_filter)

    # Revenue by method for summary
    revenue_by_method = payments.values('method').annotate(
        total=Sum('amount'),
        count=Count('id')
    )
    total_revenue = payments.aggregate(Sum('amount'))['amount__sum'] or 0
    total_bookings = payments.count()

    # ===== EXCEL (.xlsx) EXPORT =====
    if export_format == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = 'Revenue Report'

            # Styles
            header_font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
            header_fill = PatternFill(start_color='3B7A8C', end_color='3B7A8C', fill_type='solid')
            section_font = Font(name='Calibri', bold=True, size=11, color='1a3a42')
            section_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
            data_font = Font(name='Calibri', size=10)
            money_font = Font(name='Calibri', size=10, color='28a745')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title row
            ws.merge_cells('A1:M1')
            ws['A1'] = f'PickleSphere Revenue Report ({date_from} to {date_to})'
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 35

            # Summary section
            row = 3
            ws[f'A{row}'] = 'SUMMARY METRICS'
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            ws[f'A{row}'] = 'Total Revenue'
            ws[f'A{row}'].font = data_font
            ws[f'B{row}'] = f'₱{total_revenue:,.2f}'
            ws[f'B{row}'].font = money_font
            row += 1
            ws[f'A{row}'] = 'Total Paid Bookings'
            ws[f'A{row}'].font = data_font
            ws[f'B{row}'] = total_bookings
            ws[f'B{row}'].font = data_font
            row += 2

            # Revenue by Method
            ws[f'A{row}'] = 'REVENUE BY PAYMENT METHOD'
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            headers = ['Method', 'Transactions', 'Revenue']
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=h)
                cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
                cell.fill = PatternFill(start_color='6c757d', end_color='6c757d', fill_type='solid')
                cell.border = thin_border
            row += 1
            for item in revenue_by_method:
                ws.cell(row=row, column=1, value=item['method'] or 'N/A').font = data_font
                ws.cell(row=row, column=2, value=item['count']).font = data_font
                ws.cell(row=row, column=3, value=f'₱{item["total"]:,.2f}').font = money_font
                for c in range(1, 4):
                    ws.cell(row=row, column=c).border = thin_border
                row += 1
            row += 1

            # Booking Details
            ws[f'A{row}'] = 'BOOKING DETAILS'
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = section_fill
            ws.merge_cells(f'A{row}:M{row}')
            row += 1
            detail_headers = [
                'Booking ID', 'Date', 'Start Time', 'End Time', 'Court',
                'Player Name', 'Duration (hrs)', 'Equipment Fee',
                'Court Fee', 'Total Amount', 'Method', 'Transaction ID', 'Status'
            ]
            for col_idx, h in enumerate(detail_headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=h)
                cell.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
                cell.fill = PatternFill(start_color='6c757d', end_color='6c757d', fill_type='solid')
                cell.border = thin_border
            row += 1
            for payment in payments:
                res = payment.reservation
                data = [
                    res.id,
                    res.date.isoformat() if hasattr(res.date, 'isoformat') else str(res.date),
                    str(res.start_time)[:5] if res.start_time else '',
                    str(res.end_time)[:5] if res.end_time else '',
                    res.court.name if res.court else 'N/A',
                    res.user.get_full_name() or res.user.username,
                    float(res.duration_hours) if res.duration_hours else 0,
                    float(res.equipment_fee or 0),
                    float(res.subtotal or 0),
                    float(payment.amount or 0),
                    payment.method or 'N/A',
                    payment.transaction_id or 'N/A',
                    payment.status
                ]
                for col_idx, val in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    if col_idx in (8, 9, 10):  # Money columns
                        cell.number_format = '#,##0.00'
                row += 1

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 3, 30)

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="revenue_report_{date_from}_to_{date_to}.xlsx"'
            wb.save(response)
            return response

        except ImportError:
            # Fallback to CSV if openpyxl not installed
            pass

    # ===== CSV EXPORT (default) =====
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="revenue_report_{date_from}_to_{date_to}.csv"'

    writer = csv.writer(response)

    # Write header
    writer.writerow([
        'Revenue Report',
        f'From: {date_from}',
        f'To: {date_to}',
        ''
    ])
    writer.writerow([])

    # Summary Section
    writer.writerow(['SUMMARY METRICS'])
    writer.writerow(['Total Revenue', f'₱{total_revenue:.2f}'])
    writer.writerow(['Total Bookings', total_bookings])
    writer.writerow([])

    # Booking Details
    writer.writerow(['BOOKING DETAILS'])
    writer.writerow([
        'Booking ID', 'Date', 'Start Time', 'End Time', 'Court',
        'Player Name', 'Duration (hrs)', 'Equipment Fee',
        'Court Fee', 'Total Amount', 'Payment Method', 'Transaction ID', 'Status'
    ])

    for payment in payments:
        reservation = payment.reservation
        writer.writerow([
            reservation.id,
            reservation.date,
            reservation.start_time,
            reservation.end_time,
            reservation.court.name if reservation.court else 'N/A',
            reservation.user.get_full_name() or reservation.user.username,
            reservation.duration_hours,
            reservation.equipment_fee,
            reservation.subtotal,
            payment.amount,
            payment.method or 'N/A',
            payment.transaction_id or 'N/A',
            payment.status
        ])

    writer.writerow([])

    # Revenue by Method
    writer.writerow(['REVENUE BY PAYMENT METHOD'])
    writer.writerow(['Method', 'Transactions', 'Revenue'])
    for item in revenue_by_method:
        writer.writerow([
            item['method'] or 'N/A',
            item['count'],
            f"₱{item['total']:.2f}"
        ])

    return response
